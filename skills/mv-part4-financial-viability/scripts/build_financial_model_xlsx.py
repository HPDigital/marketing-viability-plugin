#!/usr/bin/env python3
"""
build_financial_model_xlsx.py
Genera el modelo financiero proyectado en formato XLSX con hojas separadas:
  - Supuestos
  - Ingresos por escenario
  - Costos
  - CAPEX
  - Flujo de caja
  - Indicadores
  - Sensibilidad

Uso:
    python build_financial_model_xlsx.py --inputs handoff_inputs.yaml --output financial_model.xlsx
"""

import argparse
import sys
from pathlib import Path

try:
    import yaml
except ImportError:
    sys.stderr.write("Falta pyyaml.\n")
    sys.exit(2)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("Falta openpyxl. Instalar con: pip install openpyxl\n")
    sys.exit(2)


HEADER_FILL = PatternFill(start_color="0B3A6E", end_color="0B3A6E", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color="0B3A6E")
THIN_BORDER = Border(
    left=Side(style="thin", color="CFCFCF"),
    right=Side(style="thin", color="CFCFCF"),
    top=Side(style="thin", color="CFCFCF"),
    bottom=Side(style="thin", color="CFCFCF"),
)


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


def style_section(cell):
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = THIN_BORDER


def autosize(ws, max_width=42):
    for col in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(length + 3, max_width)


def sheet_assumptions(wb, data):
    ws = wb.create_sheet("Supuestos")
    fm = data.get("financial_model", {})
    rows = [
        ("Caso", data.get("case_name", "")),
        ("Fecha", data.get("date", "")),
        ("Horizonte (años)", fm.get("horizon_years", 5)),
        ("Granularidad año 1", fm.get("granularity", {}).get("year_1", "monthly")),
        ("Granularidad años 2-3", fm.get("granularity", {}).get("years_2_3", "quarterly")),
        ("Granularidad años 4-5", fm.get("granularity", {}).get("years_4_5", "semestral")),
        ("Inflación anual (%)", fm.get("macro_assumptions", {}).get("inflation_annual", "")),
        ("Tipo de cambio", fm.get("macro_assumptions", {}).get("exchange_rate", "")),
        ("Tasa de descuento (%)", fm.get("macro_assumptions", {}).get("discount_rate", "")),
    ]
    h1 = ws.cell(row=1, column=1, value="Parámetro")
    h2 = ws.cell(row=1, column=2, value="Valor")
    style_header(h1)
    style_header(h2)
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k).border = THIN_BORDER
        ws.cell(row=i, column=2, value=v).border = THIN_BORDER
    autosize(ws)


def sheet_revenue(wb, data):
    ws = wb.create_sheet("Ingresos")
    rev = data.get("revenue_projection", {})
    total = rev.get("total_by_year", {})
    if not total:
        ws.cell(row=1, column=1, value="No hay datos de ingresos.")
        return

    years = sorted({i for vals in total.values() for i, _ in enumerate(vals, start=1)})
    headers = ["Escenario"] + [f"Año {y}" for y in years] + ["Total"]
    for j, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=j, value=h))

    for i, (scenario, vals) in enumerate(total.items(), start=2):
        ws.cell(row=i, column=1, value=scenario.capitalize()).font = Font(bold=True)
        for j, v in enumerate(vals, start=2):
            ws.cell(row=i, column=j, value=v).number_format = "#,##0"
        ws.cell(row=i, column=len(years) + 2, value=sum(vals)).number_format = "#,##0"
        for j in range(1, len(headers) + 1):
            ws.cell(row=i, column=j).border = THIN_BORDER
    autosize(ws)


def sheet_costs(wb, data):
    ws = wb.create_sheet("Costos")
    cs = data.get("cost_structure", {})
    headers = ["Categoría / Año"] + [f"Año {i+1}" for i in range(5)]
    for j, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=j, value=h))

    blocks = [
        ("Costos variables (base)", cs.get("variable_costs_by_year", {}).get("base", [])),
        ("Costos fijos operativos", cs.get("fixed_costs_operational_by_year", [])),
        ("Costos comerciales y mkt", cs.get("fixed_costs_commercial_marketing_by_year", [])),
        ("Estructura corporativa", cs.get("corporate_structure_by_year", [])),
        ("Costos financieros", cs.get("financial_costs_by_year", [])),
        ("Ahorros Blue Ocean (–)", cs.get("blue_ocean_savings_by_year", [])),
    ]
    for i, (label, vals) in enumerate(blocks, start=2):
        cell = ws.cell(row=i, column=1, value=label)
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        for j, v in enumerate(vals, start=2):
            c = ws.cell(row=i, column=j, value=v)
            c.number_format = "#,##0"
            c.border = THIN_BORDER
    autosize(ws)


def sheet_capex(wb, data):
    ws = wb.create_sheet("CAPEX")
    cp = data.get("capex_plan", {})
    initial = cp.get("initial_capex", {})
    expansion = cp.get("expansion_capex_by_year", [])

    style_header(ws.cell(row=1, column=1, value="CAPEX inicial"))
    style_header(ws.cell(row=1, column=2, value="USD"))
    rows_initial = [("Taller", initial.get("workshop", 0)),
                    ("Stock", initial.get("stock", 0)),
                    ("Regulatorio", initial.get("regulatory", 0)),
                    ("Comercial", initial.get("commercial", 0)),
                    ("TOTAL", initial.get("total", sum(initial.get(k, 0) for k in ["workshop", "stock", "regulatory", "commercial"])))]
    for i, (k, v) in enumerate(rows_initial, start=2):
        c1 = ws.cell(row=i, column=1, value=k)
        c2 = ws.cell(row=i, column=2, value=v)
        c2.number_format = "#,##0"
        if k == "TOTAL":
            c1.font = Font(bold=True)
            c2.font = Font(bold=True)
        c1.border = THIN_BORDER
        c2.border = THIN_BORDER

    row_offset = len(rows_initial) + 4
    style_header(ws.cell(row=row_offset, column=1, value="CAPEX expansión"))
    for j, y in enumerate(range(1, 6), start=2):
        style_header(ws.cell(row=row_offset, column=j, value=f"Año {y}"))
    ws.cell(row=row_offset + 1, column=1, value="USD").font = Font(bold=True)
    for j, v in enumerate(expansion, start=2):
        c = ws.cell(row=row_offset + 1, column=j, value=v)
        c.number_format = "#,##0"
        c.border = THIN_BORDER

    wc_row = row_offset + 3
    ws.cell(row=wc_row, column=1, value="Capital de trabajo permanente").font = Font(bold=True)
    c_wc = ws.cell(row=wc_row, column=2, value=cp.get("working_capital_permanent", 0))
    c_wc.number_format = "#,##0"
    autosize(ws)


def sheet_cash_flow(wb, data):
    ws = wb.create_sheet("Flujo de caja")
    cf = data.get("cash_flow", {})
    headers = ["Concepto / Año"] + [f"Año {i+1}" for i in range(5)]
    for j, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=j, value=h))

    blocks = [
        ("FCF anual", cf.get("free_cash_flow_by_year", [])),
        ("FCF acumulado", cf.get("cumulative_fcf_by_year", [])),
        ("CF inversión", cf.get("investment_cf_by_year", [])),
        ("CF financiación", cf.get("financing_cf_by_year", [])),
    ]
    for i, (label, vals) in enumerate(blocks, start=2):
        c1 = ws.cell(row=i, column=1, value=label)
        c1.font = Font(bold=True)
        c1.border = THIN_BORDER
        for j, v in enumerate(vals, start=2):
            c = ws.cell(row=i, column=j, value=v)
            c.number_format = "#,##0;[Red]-#,##0"
            c.border = THIN_BORDER

    if "cash_inflection_month" in cf:
        ws.cell(row=len(blocks) + 4, column=1, value="Punto de inflexión de caja (mes)").font = Font(bold=True)
        ws.cell(row=len(blocks) + 4, column=2, value=cf["cash_inflection_month"])
    autosize(ws)


def sheet_indicators(wb, data):
    ws = wb.create_sheet("Indicadores")
    vi = data.get("viability_indicators", {})
    ue = data.get("unit_economics", {})

    style_header(ws.cell(row=1, column=1, value="Indicador"))
    style_header(ws.cell(row=1, column=2, value="Pesimista"))
    style_header(ws.cell(row=1, column=3, value="Base"))
    style_header(ws.cell(row=1, column=4, value="Optimista"))

    npv = vi.get("npv", {})
    irr = vi.get("irr", {})
    pb = vi.get("payback_simple_months", {})
    pbd = vi.get("payback_discounted_months", {})
    rows = [
        ("VAN", npv.get("pessimistic"), npv.get("base"), npv.get("optimistic")),
        ("TIR (%)", irr.get("pessimistic"), irr.get("base"), irr.get("optimistic")),
        ("Payback simple (meses)", pb.get("pessimistic"), pb.get("base"), pb.get("optimistic")),
        ("Payback descontado (meses)", pbd.get("pessimistic"), pbd.get("base"), pbd.get("optimistic")),
    ]
    for i, row in enumerate(rows, start=2):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN_BORDER
            if j == 1:
                c.font = Font(bold=True)
            elif isinstance(v, (int, float)):
                c.number_format = "#,##0"

    row_ue = len(rows) + 4
    style_header(ws.cell(row=row_ue, column=1, value="Métricas de unidad económica"))
    style_header(ws.cell(row=row_ue, column=2, value="Valor"))
    ue_rows = [
        ("CAC blended", ue.get("cac_blended")),
        ("LTV blended", ue.get("ltv_blended")),
        ("LTV / CAC", ue.get("ltv_cac_ratio")),
        ("Payback CAC (meses)", ue.get("payback_cac_months")),
        ("Margen contributivo unidad", ue.get("contribution_margin_per_unit")),
    ]
    for i, (k, v) in enumerate(ue_rows, start=row_ue + 1):
        c1 = ws.cell(row=i, column=1, value=k)
        c1.font = Font(bold=True)
        c1.border = THIN_BORDER
        c2 = ws.cell(row=i, column=2, value=v)
        c2.number_format = "#,##0.0" if isinstance(v, float) else "#,##0"
        c2.border = THIN_BORDER

    autosize(ws)


def sheet_sensitivity(wb, data):
    ws = wb.create_sheet("Sensibilidad")
    sa = data.get("sensitivity_analysis", {})
    tornado = sa.get("tornado_data", [])

    style_header(ws.cell(row=1, column=1, value="Variable"))
    style_header(ws.cell(row=1, column=2, value="Impacto −20%"))
    style_header(ws.cell(row=1, column=3, value="Impacto +20%"))
    for i, item in enumerate(tornado, start=2):
        ws.cell(row=i, column=1, value=item.get("variable", "")).border = THIN_BORDER
        c2 = ws.cell(row=i, column=2, value=item.get("impact_low", 0))
        c3 = ws.cell(row=i, column=3, value=item.get("impact_high", 0))
        c2.number_format = "#,##0;[Red]-#,##0"
        c3.number_format = "#,##0;[Red]-#,##0"
        c2.border = THIN_BORDER
        c3.border = THIN_BORDER

    autosize(ws)


def sheet_verdict(wb, data):
    ws = wb.create_sheet("Veredicto")
    fv = data.get("final_verdict", {})
    rows = [
        ("Veredicto", fv.get("veredict", "—")),
        ("Capital total necesario", fv.get("capital_needed_total", "—")),
    ]
    style_header(ws.cell(row=1, column=1, value="Concepto"))
    style_header(ws.cell(row=1, column=2, value="Valor"))
    for i, (k, v) in enumerate(rows, start=2):
        c1 = ws.cell(row=i, column=1, value=k)
        c1.font = Font(bold=True)
        c1.border = THIN_BORDER
        c2 = ws.cell(row=i, column=2, value=v)
        c2.border = THIN_BORDER

    cs = fv.get("capital_structure_proposed", {})
    if cs:
        row = len(rows) + 4
        style_header(ws.cell(row=row, column=1, value="Estructura de capital"))
        style_header(ws.cell(row=row, column=2, value="USD"))
        for i, (k, v) in enumerate(cs.items(), start=row + 1):
            ws.cell(row=i, column=1, value=k.capitalize()).border = THIN_BORDER
            c2 = ws.cell(row=i, column=2, value=v)
            c2.number_format = "#,##0"
            c2.border = THIN_BORDER

    cond = fv.get("conditional_on", [])
    if cond:
        row = ws.max_row + 3
        style_header(ws.cell(row=row, column=1, value="Condicional a"))
        for i, c in enumerate(cond, start=row + 1):
            ws.cell(row=i, column=1, value=f"• {c}").border = THIN_BORDER

    autosize(ws)


def main():
    parser = argparse.ArgumentParser(description="Construir el modelo financiero XLSX.")
    parser.add_argument("--inputs", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    if not args.inputs.exists():
        sys.stderr.write(f"No se encuentra el archivo de inputs: {args.inputs}\n")
        sys.exit(1)

    with open(args.inputs, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}

    wb = Workbook()
    wb.remove(wb.active)
    sheet_assumptions(wb, data)
    sheet_revenue(wb, data)
    sheet_costs(wb, data)
    sheet_capex(wb, data)
    sheet_cash_flow(wb, data)
    sheet_indicators(wb, data)
    sheet_sensitivity(wb, data)
    sheet_verdict(wb, data)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Modelo financiero generado en {args.output}")


if __name__ == "__main__":
    main()
